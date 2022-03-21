import openpyxl

group_cell = dict()


# init - просматривает все xlsx файлы и создает словарь с нахождением расписания этих групп
def init():
    for number_of_xlxs in range(1, 4):
        sheet = openpyxl.load_workbook("iit" + str(number_of_xlxs) + ".xlsx").active
        for column in range(5, sheet.max_column, 5):
            if sheet[2][column].value != "День недели":
                group_cell[sheet[2][column].value] = column


# get_xlsx - по имени группы выводит название файла, в котором лежит расписание с этой группой
def get_xlsx(group_name):
    if group_name[-2:] == "21":
        return "iit1.xlsx"
    elif group_name[-2:] == "20":
        return "iit2.xlsx"
    elif group_name[-2:] == "19":
        return "iit3.xlsx"
    elif group_name[-2:] == "18":
        return "iit4.xlsx"
    else:
        return None


def print_week(group_name, week):
    if group_cell.get(group_name) is not None:
        excel_filename = get_xlsx(group_name)
        sheet = openpyxl.load_workbook(excel_filename).active

        for day in range(4, 64, 12):
            print(sheet[day][0].value)
            for para in range(day + week, day + 12, 2):
                if sheet[para][group_cell[group_name]].value is not None:
                    if sheet[para][group_cell[group_name] - 3].value is not None:
                        print(str(sheet[para][group_cell[group_name] - 3].value), end="\t")
                        print(str(sheet[para][group_cell[group_name] - 2].value), end="\t")
                    else:
                        print(str(sheet[para - 1][group_cell[group_name] - 3].value), end="\t")
                        print(str(sheet[para - 1][group_cell[group_name] - 2].value), end="\t")

                    print(str(sheet[para][group_cell[group_name]].value))
            print()
    else:
        return "Ваша группа не найдена."

init()
print_week("ИАБО-01-20", 1)