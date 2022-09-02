"""Работа с .xlsx(excel) файлами расписания РТУ МИРЭА через модуль openpyxl"""
import os

import openpyxl

import soup_worker

# Словаря для хранения пар {группа}:{[файл], [колонка]}
group_cell = dict()


# init() - вызов parse_mirea() и составление словаря
def __init__():
    """Инициализация Хэш-таблицы групп

    Находит название группы в таблице-excel, запоминает
     столбец и файл, в котором был найден столбец"""
    soup_worker.parse_mirea()

    files = [f for f in os.listdir('.') if os.path.isfile(f)]
    for file in files:
        if file.endswith(".xlsx"):
            workbook = openpyxl.load_workbook(file)
            for sheet in workbook:
                for column in range(5, sheet.max_column, 5):
                    if sheet[2][column].value != "День недели":
                        group_cell[sheet[2][column].value] = [file, column, sheet.title]


# is_group_exists() - проверка существования группы в словаре
def is_group_exists(group_name):
    """Проверка на существование группы

    Проверка, есть ли ключ названия группы в данной
     Хэш-таблице"""
    if group_cell.get(group_name.upper()) is not None:
        return True
    return False


# print_week() - вывод недели расписания по имени группы (week: 0 - нечетная, 1 - четная)
def print_week(group_name, week):
    """Вывод недели расписания по имени группы и четности недели

    Имея номер колонки excel-таблицы с определенной группой,
     происходит разбор таблицы и вычленение из нее важных ячеек с
      названием предмета, номером пары, временем, кабинетом"""
    group_name = group_name.upper()

    if group_cell.get(group_name)[1] is not None:
        answer = "```\n"
        if week == 0:
            answer += '⚡Нечетная неделя\n\n'
        else:
            answer += '⚡Четная неделя\n\n'

        excel_filename = group_cell.get(group_name)[0]
        workbook = openpyxl.load_workbook(excel_filename)
        sheet = workbook[group_cell.get(group_name)[2]]

        for day in range(4, 76, 12):
            answer += "🔥" + sheet[day][0].value + '\n'
            for para in range(day + week, day + 12, 2):
                if sheet[para][group_cell[group_name][1]].value != "":
                    if sheet[para][1].value is not None:
                        answer += '№' + str(sheet[para][1].value) + '\t'
                        answer += str(sheet[para][2].value) + '\t'
                        answer += str(sheet[para][3].value) + '\t'
                    else:
                        answer += '№' + str(sheet[para - 1][1].value) + '\t'
                        answer += str(sheet[para - 1][2].value) + '\t'
                        answer += str(sheet[para - 1][3].value) + '\t'

                    answer += str(sheet[para][group_cell[group_name][1] + 3].value).replace('\n', '/ ') + '\n'
                    if sheet[para][group_cell[group_name][1]].value is not None:
                        answer += str(sheet[para][group_cell[group_name][1]].value) + '\n\n'
                    else:
                        n = para
                        while sheet[n][group_cell[group_name][1]].value is None:
                            n -= 1
                        answer += str(sheet[n][group_cell[group_name][1]].value) + '\n\n'
            answer += '\n\n'

        answer += "```"
        return answer
    else:
        return "Ваша группа не найдена."


# print_day() - вывод дня расписания по имени группы для текущей недели (1 - понедельник, 2 - вторник ...)
def print_day(group_name, week, day):
    """Вывод расписания дня по имени группы, четности недели и номером дня

    Аналогичная работа, как и у функции print_week(), но
     в данной функции учитывается и день недели (1 - ПН, ..., 7 - ВС)"""
    group_name = group_name.upper()

    if group_cell.get(group_name)[1] is not None:
        answer = "```\n"
        excel_filename = group_cell.get(group_name)[0]
        workbook = openpyxl.load_workbook(excel_filename)
        sheet = workbook[group_cell.get(group_name)[2]]

        answer += "🔥" + sheet[4 + 12 * (day - 1)][0].value + '\n'
        for para in range(4 + 12 * (day - 1) + week, 16 + 12 * (day - 1), 2):
            if sheet[para][group_cell[group_name][1]].value != "":
                if sheet[para][1].value is not None:
                    answer += '№' + str(sheet[para][1].value) + '\t'
                    answer += str(sheet[para][2].value) + '\t'
                    answer += str(sheet[para][3].value) + '\t'
                else:
                    answer += '№' + str(sheet[para - 1][1].value) + '\t'
                    answer += str(sheet[para - 1][2].value) + '\t'
                    answer += str(sheet[para - 1][3].value) + '\t'

                answer += str(sheet[para][group_cell[group_name][1] + 3].value).replace('\n', '/ ') + '\n'
                if sheet[para][group_cell[group_name][1]].value is not None:
                    answer += str(sheet[para][group_cell[group_name][1]].value) + '\n\n'
                else:
                    n = para
                    while sheet[n][group_cell[group_name][1]].value is None:
                        n -= 1
                    answer += str(sheet[n][group_cell[group_name][1]].value) + '\n\n'

        answer += "```"
        return answer
    else:
        return "Ваша группа не найдена."
